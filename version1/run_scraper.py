"""
Tender Aggregator — All Portals Combined
==========================================
Runs scrapers across multiple procurement portals,
collects open tenders with DIRECT document download links,
and saves a single Excel workbook sorted by closing date.

Usage:
  # Scrape a specific portal:
  python run_scraper.py --portals gem eprocure --from 01/01/2024 --to 31/03/2024

  # Scrape all portals:
  python run_scraper.py --portals all --from 01/06/2025 --to 30/06/2025

  # Just generate a document link (no scraping):
  python run_scraper.py --make_link --portal gem --id 9032455

  # List all supported portals:
  python run_scraper.py --list_portals
"""

import argparse
import logging
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from portal_adapters import ALL_PORTALS, PORTAL_ALIASES, make_doc_link, Tender

try:
    from eprocure_scraper import make_driver
except ImportError:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    def make_driver(headless=False):
        opts = Options()
        if headless:
            opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        return webdriver.Chrome(options=opts)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────
# Date parsing
# ─────────────────────────────────────────────────────────────────
def parse_date_for_sort(s: str):
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return datetime.max


# ─────────────────────────────────────────────────────────────────
# Excel export with formatting
# ─────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")   # dark blue
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL     = PatternFill("solid", fgColor="EBF3FB")   # light blue
LINK_FONT    = Font(color="0563C1", underline="single")


def export_to_excel(tenders: list[Tender], output_path: str):
    if not tenders:
        log.warning("No tenders to export.")
        return

    rows = [t.to_dict() for t in tenders]
    df = pd.DataFrame(rows)

    # Sort by closing date
    df["_sort"] = df["Closing Date"].apply(parse_date_for_sort)
    df = df.sort_values("_sort").drop(columns=["_sort"])
    df = df.reset_index(drop=True)

    # Column order
    col_order = [
        "Closing Date",
        "Closing Date (from doc)",
        "Published Date",
        "State",
        "Portal",
        "Organisation",
        "Tender Title",
        "Tender Ref No",
        "NIT Download Link",
        "BOQ Download Link",
        "Other Doc Links",
        "Opening Date",
        "Detail Page URL",
        "Tender ID",
    ]
    existing = [c for c in col_order if c in df.columns]
    extras   = [c for c in df.columns if c not in existing]
    df = df[existing + extras]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="All Tenders")
        ws = writer.sheets["All Tenders"]

        # Style header row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows + hyperlinks
        link_cols = {
            col: idx + 1
            for idx, col in enumerate(df.columns)
            if "Link" in col or col in ("Detail Page URL",)
        }

        for row_idx in range(2, len(df) + 2):
            # Alternating row color
            fill = ALT_FILL if row_idx % 2 == 0 else None

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if fill:
                    cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Hyperlink styling for link columns
            for col_name, col_idx in link_cols.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                url = str(cell.value or "").strip()
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = LINK_FONT
                    cell.value = "📄 Download" if "Download" in col_name else "🔗 Link"

        # Column widths
        col_widths = {
            "Closing Date": 14,
            "Closing Date (from doc)": 18,
            "Published Date": 14,
            "State": 16,
            "Portal": 28,
            "Organisation": 30,
            "Tender Title": 45,
            "Tender Ref No": 22,
            "NIT Download Link": 16,
            "BOQ Download Link": 16,
            "Other Doc Links": 16,
            "Opening Date": 14,
            "Detail Page URL": 16,
            "Tender ID": 22,
        }
        for col_idx, col_name in enumerate(df.columns, start=1):
            w = col_widths.get(col_name, 20)
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add summary sheet
        summary_data = df.groupby("State")["Tender Title"].count().reset_index()
        summary_data.columns = ["State/Portal", "Tender Count"]
        summary_data = summary_data.sort_values("Tender Count", ascending=False)
        summary_data.to_excel(writer, sheet_name="Summary", index=False)

    log.info(f"✓ Exported {len(df)} tenders → {output_path}")
    return df


# ─────────────────────────────────────────────────────────────────
# Main runner
# ─────────────────────────────────────────────────────────────────
def run_portals(
    portal_names: list[str],
    from_date: str,
    to_date: str,
    headless: bool = False,
    output: str = None,
) -> list[Tender]:

    # Resolve "all"
    if "all" in portal_names:
        portal_names = list(ALL_PORTALS.keys())

    # Resolve aliases
    resolved = []
    for p in portal_names:
        key = PORTAL_ALIASES.get(p.lower(), p.lower())
        if key in ALL_PORTALS:
            resolved.append(key)
        else:
            log.warning(f"Unknown portal: {p} — skipping")

    if not resolved:
        log.error("No valid portals specified.")
        return []

    log.info(f"Portals to scrape: {resolved}")
    log.info(f"Date range: {from_date} → {to_date}")

    driver = make_driver(headless=headless)
    all_tenders: list[Tender] = []

    try:
        for portal_key in resolved:
            adapter = ALL_PORTALS[portal_key]
            log.info(f"\n{'='*60}")
            log.info(f"Scraping: {adapter.name}")
            log.info(f"{'='*60}")
            try:
                tenders = adapter.search_tenders(driver, from_date, to_date)
                log.info(f"  → {len(tenders)} tenders found")
                all_tenders.extend(tenders)
            except Exception as e:
                log.error(f"  ERROR scraping {adapter.name}: {e}")
            time.sleep(2)  # polite pause between portals
    finally:
        driver.quit()

    if not output:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        portals_str = "_".join(resolved[:3])
        output = f"tenders_{portals_str}_{ts}.xlsx"

    export_to_excel(all_tenders, output)
    return all_tenders


# ─────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Aggregate open tenders from Indian govt procurement portals"
    )
    parser.add_argument(
        "--portals", nargs="+", default=["eprocure"],
        help="Portal keys to scrape (e.g. gem eprocure maharashtra) or 'all'"
    )
    parser.add_argument("--from", dest="from_date", default="01/01/2024",
                        help="Start published date DD/MM/YYYY")
    parser.add_argument("--to",   dest="to_date",   default="31/12/2024",
                        help="End published date DD/MM/YYYY")
    parser.add_argument("--headless", action="store_true",
                        help="Run Chrome in headless mode (no GUI)")
    parser.add_argument("--output", default=None,
                        help="Output Excel filename")
    parser.add_argument("--make_link", action="store_true",
                        help="Just generate a document link (no scraping)")
    parser.add_argument("--portal", default="gem",
                        help="Portal name for --make_link")
    parser.add_argument("--id", dest="doc_id", default="",
                        help="Document/tender ID for --make_link")
    parser.add_argument("--doc_type", default="NIT",
                        help="Document type: NIT or BOQ")
    parser.add_argument("--list_portals", action="store_true",
                        help="List all supported portals and exit")
    args = parser.parse_args()

    # List portals
    if args.list_portals:
        print("\nSupported portals:\n")
        for key, adapter in ALL_PORTALS.items():
            print(f"  {key:<20} {adapter.name}")
        print(f"\nAliases: {PORTAL_ALIASES}")
        sys.exit(0)

    # Quick link generation
    if args.make_link:
        if not args.doc_id:
            print("Error: --id is required with --make_link")
            sys.exit(1)
        link = make_doc_link(args.portal, args.doc_id, args.doc_type)
        print(f"\nDirect {args.doc_type} download link for {args.portal}/{args.doc_id}:")
        print(f"  {link}\n")
        sys.exit(0)

    # Run full scraper
    run_portals(
        portal_names=args.portals,
        from_date=args.from_date,
        to_date=args.to_date,
        headless=args.headless,
        output=args.output,
    )


if __name__ == "__main__":
    main()
